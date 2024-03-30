# Import generales
try:
    import flask, os, io
    from flask import Flask, request, jsonify, render_template, send_from_directory
    from dotenv import load_dotenv
    from flask_cors import CORS
    from sharepy import connect
    import pandas as pd
    import openpyxl as px

    # Import de modelos
    from sql.db import init_db, db
    from sql.models.Users import Users
    from sql.models.Accessories import Accessories
    from sql.models.Computers import Computadoras
    from sql.models.UserComputers import UserComputer

    from sqlalchemy.sql import func, exists, and_, or_
    from sqlalchemy.exc import IntegrityError

except Exception as e:
    print(f"Error: {e}")

# Configuración de la app
app = Flask(__name__)
cors = CORS(app, resources={r"/api/*": {"origins": "*"}})

# Configuración de la base de datos
load_dotenv('.env')

app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('SQLALCHEMY_DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')

# Inicialización de la base de datos
db = init_db(app)

def connectToExcelSharepoint():
    try:
        # creando const para la conexion con sharepoint
        URL = os.getenv('URL') # URL del excel en sharepoint
        SHAREPOINT_USER = os.getenv('SHAREPOINT_USER')
        SHAREPOINT_PASSWORD = os.getenv('SHAREPOINT_PASSWORD')

        s = connect(URL, SHAREPOINT_USER, SHAREPOINT_PASSWORD)
        return s
    except Exception as e:
        print(f"Error: {e}")
        return None
    
def onUserBlock(user):
    '''
    Cuando un usuario se bloquea en el sistema
    se buscara en el excel la casilla con el mail del usuario y se pintara de rojo la fila
    '''
    try:
        # conectando a sharepoint
        excel = connectToExcelSharepoint()

        # si no se puede conectar a sharepoint
        if not excel:
            return False
                
        # obteniendo el archivo de excel
        file = excel.getfile()

        with io.BytesIO(file.content) as f:
            # leyendo el archivo de excel
            wb = px.load_workbook(f)
            ws = wb.active

            # buscando el usuario en el excel
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
                for cell in row:
                    if cell.value == user.mail:
                        # pintando la fila de rojo
                        for cell in row:
                            cell.fill = px.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                        break
                    else:
                        print(f"Usuario no encontrado en el excel")
                    
            # guardando los cambios en el excel
            wb.save('file.xlsx')
            excel.upload('file.xlsx')

            return True
        
    except Exception as e:
        print(f"Error: {e}")
        return False

# Rutas
@app.route('/api/v1.0/users', methods=['GET'])
def users():
    try:
        users = Users.query.all()
        return jsonify([user.serialize() for user in users])
    except:
        return jsonify({'message': 'Error al obtener los usuarios'})

# search user by name
@app.route('/api/v1.0/users/<string:nombre>', methods=['GET'])
def search_user(nombre):
    try:
        nombre = nombre.capitalize()
        users = Users.query.filter(and_(or_(
            Users.nombre.like(f'{nombre}%'),
            Users.apellido.like(f'{nombre}%'),
            Users.mail.like(f'{nombre}%')
        ), Users.is_active == True)).all()

        if len(users) == 0:
            return jsonify({'message': 'Usuario no encontrado'})
        
        # get computer of user
        for user in users:
            user.computadora = [computer.serialize() for computer in user.computadora_id]
        
        return jsonify([user.serialize() for user in users])                  
    except:
        return jsonify({'message': 'Error al obtener los usuarios'})
    
@app.route('/api/v1.0/usersDisabled/<string:nombre>', methods=['GET'])
def search_user_disabled(nombre):
    try:
        nombre = nombre.capitalize()
        users = Users.query.filter(and_(or_(
            Users.nombre.like(f'{nombre}%'),
            Users.apellido.like(f'{nombre}%'),
            Users.mail.like(f'{nombre}%')
        ), Users.is_active == False)).all()

        if len(users) == 0:
            return jsonify({'message': 'Usuario no encontrado'})

        return jsonify([user.serialize() for user in users])
    except:
        return jsonify({'message': 'Error al obtener los usuarios'})

    
@app.route('/api/v1.0/users/<int:id>', methods=['GET'])
def user(id):
    try:
        user = Users.query.get(id)
        return jsonify(user.serialize())
    except:
        return jsonify({'message': 'Error al obtener el usuario'})

@app.route('/api/v1.0/users', methods=['POST'])
def create_user():
    try:
        data = request.get_json()
        user = Users(**data)
        
        db.session.add(user)
        
        # recibe parametro `computadora_id` para asignar la computadora al usuario
        if 'computadora_id' in data:
            computer = Computadoras.query.get(data['computadora_id'])
            if computer:
                user.computadora_id = computer
                db.session.commit()
            # si la computadora esta asignada a otro usuario, se asigna la nueva al nuevo usuario
            elif computer.user_id:
                computer = Computadoras.query.filter_by(user_id=user.id).first()
                computer.user_id = None
                user.computadora_id = computer
                db.session.commit()
            # si la computadora no existe, se crea una nueva
            else:
                computer = Computadoras(**data)
                db.session.add(computer)
                db.session.commit()
                user.computadora_id = computer
                db.session.commit() 
            
        db.session.commit()        
        
        return jsonify(user.serialize())
    except:
        return jsonify({'message': 'Error al crear el usuario'})    
    
@app.route('/api/v1.0/userDeactivate/<int:id>', methods=['POST'])
def deactivate_user(id):
    try:
        user = Users.query.get(id)

        if not user:
            return jsonify({'message': 'Usuario no encontrado'}), 404
        
        if user.is_active == False:
            return jsonify({'message': 'El usuario ya se encuentra desactivado'}), 400
        
        nombre = user.nombre

        # get accessories
        try:
            accessories = Accessories.query.filter_by(workday_id=user.workday_id).all()
            for accessory in accessories:
                db.session.delete(accessory)
                db.session.commit()
        except:
            return jsonify({'message': 'Error al eliminar los accesorios'})
        
        # get computers
        try:
            computers = UserComputer.query.filter_by(user_id=user.id).all()
            for computer in computers:
                db.session.delete(computer)
                db.session.commit()
        except:
            return jsonify({'message': 'Error al eliminar las computadoras'})
        
        # bloquear usuario en excel
        if onUserBlock(user):
            print(f"Usuario {nombre} bloqueado en el excel")
        else:
            print(f"Error al bloquear usuario {nombre} en el excel")
        
        user.is_active = False
        user.updated_at = func.now()
        db.session.commit()
        
        return jsonify({'message': f'Usuario {nombre} desactivado'})
    except:
        return jsonify({'message': 'Error al desactivar el usuario'})
    
@app.route('/api/v1.0/userActivate/<int:id>', methods=['POST'])
def activate_user(id):
    try:
        user = Users.query.get(id)

        if not user:
            return jsonify({'message': 'Usuario no encontrado'}), 404
        
        nombre = user.nombre

        user.is_active = True
        user.updated_at = func.now()
        db.session.commit()
        
        return jsonify({'message': f'Usuario {nombre} activado'})
    except:
        return jsonify({'message': 'Error al activar el usuario'})

@app.route('/api/v1.0/users/<int:id>', methods=['PUT'])
def update_user(id):
    try:
        # Buscar al usuario por ID
        user = Users.query.get(id)
        
        if not user:
            return jsonify({'message': 'Usuario no encontrado'}), 404
        
        data = request.get_json()

        # Actualizar los atributos del usuario con los nuevos datos si están presentes
        user.nombre = data.get('nombre', user.nombre)
        user.apellido = data.get('apellido', user.apellido)
        user.mail = data.get('mail', user.mail)
        user.usuario = data.get('usuario', user.usuario)

        # Obtener los IDs de las computadoras asignadas del cuerpo de la solicitud
        computadoras_ids = data.get('computadoras_ids', [])

        # Limpiar asignaciones de computadoras que no están en la lista
        UserComputer.query.filter(UserComputer.user_id == id, ~UserComputer.computer_id.in_(computadoras_ids)).delete(synchronize_session=False)

        for computadora_id in computadoras_ids:
            # Buscar la computadora por ID
            computadora = Computadoras.query.get(computadora_id)

            if not computadora:
                return jsonify({'message': f'Computadora con ID {computadora_id} no encontrada'}), 404
                
            # Verificar si la computadora ya está asignada a otro usuario
            existing_assignment = UserComputer.query.filter_by(computer_id=computadora_id).first()
            if existing_assignment and existing_assignment.user_id != id:
                assigned_user = Users.query.get(existing_assignment.user_id)
                return jsonify({'message': f'La computadora ya está asignada a otro usuario: {assigned_user.nombre}'}), 201
            elif existing_assignment and existing_assignment.user_id == id:
                # Si la computadora ya está asignada al usuario, solo agregar la computadora que NO está asignada
                continue

            # Asignar la computadora al usuario
            user_computer = UserComputer(user_id=id, computer_id=computadora_id)
            db.session.add(user_computer)

        # Actualizar la fecha de modificación del usuario
        user.updated_at = func.now()

        # Confirmar los cambios en la base de datos
        db.session.commit()
        
        return jsonify(user.serialize())
    except IntegrityError:
        db.session.rollback()
        return jsonify({'message': 'Error de integridad al asignar la computadora al usuario'}), 500
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': 'Error al actualizar el usuario', 'error': str(e)}), 500
    
@app.route('/api/v1.0/users/<int:id>/force', methods=['POST'])
def forceUpdateComputer(id):
    try:
        data = request.get_json()
        
        # Obtener el ID de la computadora asignada del cuerpo de la solicitud
        computadoras_ids = data.get('computadoras_ids', [])
        
        for computadora_id in computadoras_ids:
            # eliminar todos los usuarios que ya tengan asignada la computadora
            UserComputer.query.filter_by(computer_id=computadora_id).delete()
            db.session.commit()

            # si el usuario tiene mas computadoras que las que se estan asignando, borrara las que no esten en la lista
            if len(UserComputer.query.filter_by(user_id=id).all()) > len(computadoras_ids):
                for user_computer in UserComputer.query.filter_by(user_id=id).all():
                    if user_computer.computer_id not in computadoras_ids:
                        UserComputer.query.filter_by(computer_id=user_computer.computer_id).delete()
                        db.session.commit()
                    else:
                        continue
                
            # asignar la computadora al usuario
            user_computer = UserComputer(user_id=id, computer_id=computadora_id)
            db.session.add(user_computer)
            db.session.commit()
            
        return jsonify({'message': 'Computadora actualizada'})
    except:
        return jsonify({'message': 'Error al actualizar la computadora', 'error': str(e)}), 500

# Accesorios rutas
@app.route('/api/v1.0/accessories', methods=['GET'])
def accessories():
    try:
        accessories = Accessories.query.all()
        return jsonify([accessory.serialize() for accessory in accessories])
    except Exception as e:
        return jsonify({'message': 'Error al obtener los accesorios', 'error': str(e)})

@app.route('/api/v1.0/accessories/<int:workday_id>', methods=['POST'])
def create_accessory(workday_id):
    try:
        data = request.get_json()
        accessory = Accessories(workday_id=workday_id, **data)
        
        print(f"Datos recibidos: {data}")
        
        db.session.add(accessory)
        db.session.commit()
               
        return jsonify(accessory.serialize())
    except Exception as e:
        return jsonify({'message': 'Error al crear el accesorio', 'error': str(e)})

@app.route('/api/v1.0/accessories/<int:id>', methods=['PUT'])
def update_accessory(id):
    try:
        accessory = Accessories.query.get(id)
        data = request.get_json()
        accessory.update(**data)
        db.session.commit()
        return jsonify(accessory.serialize())
    except:
        return jsonify({'message': 'Error al actualizar el accesorio'}) 

@app.route('/api/v1.0/accessories/<int:id>', methods=['DELETE'])
def delete_accessory(id):
    try:
        accessory = Accessories.query.get(id)
        db.session.delete(accessory)
        db.session.commit()
        return jsonify(accessory.serialize())
    except: 
        return jsonify({'message': 'Error al eliminar el accesorio'})

@app.route('/api/v1.0/accessories/<string:workday_id>', methods=['GET'])
def accessories_by_user(workday_id):
    try:
        accessories = Accessories.query.filter_by(workday_id=workday_id).all()
        return jsonify([accessory.serialize() for accessory in accessories])
    except Exception as e:
        return jsonify({'message': 'Error al obtener los accesorios', 'error': str(e)})
    
# Computadoras rutas
@app.route('/api/v1.0/computadoras', methods=['GET'])
def get_available_computers():
    try:
        computers = Computadoras.query.all()
        return jsonify([computer.serialize() for computer in computers])
    except Exception as e:
        return jsonify({'message': 'Error al recuperar las computadoras', 'error': str(e)}), 500
    
# inicio de la app
if __name__ == '__main__':
    app.run(debug=True, port=8010)